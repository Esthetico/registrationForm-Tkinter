#-------------------------------------------------------------------------------
# Name:        Registration Form
# Purpose:     Personal practice
#
# Author:      arthur
#
# Created:     09/11/2022
# Copyright:   (c) arthur 2022
# Licence:     < UNLICENSED >
#-------------------------------------------------------------------------------
from openpyxl import *
from tkinter import *

# Global variables
theme_mode = True
dark = '#26242f'
white = '#c0c4c1'
textFg = '#42f557'

# Global declaration of wb and sheet variable
# Proving absolute path of excel file from which
# to read and write data
wb = load_workbook(r'/home/mr_unicorn/Documents/School/registration_Data.xlsx')

# Create sheet object
sheet = wb.active

# Modifying Excel spreadsheet to fit registration form purposes
def excel():

    # Resizing spreadshit columns
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50

    # Attributing values to the columns
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Course"
    sheet.cell(row=1, column=3).value = "Semester"
    sheet.cell(row=1, column=4).value = "Form Number"
    sheet.cell(row=1, column=5).value = "Contact Number"
    sheet.cell(row=1, column=6).value = "Email id"
    sheet.cell(row=1, column=7).value = "Address"

# Functions to set focus on different fields
def focus1(event):
    course_field.focus_set()

def focus2(event):
    sem_field.focus_set()

def focus3(event):
    form_no_field.focus_set()

def focus4(event):
    contact_no_field.focus_set()

def focus5(event):
    email_id_field.focus.set()

def focus6(event):
    address_field.focus_set()

# function to clear text entry box content
def clear():
    name_field.delete(0, END)
    course_field.delete(0, END)
    sem_field.delete(0, END)
    form_no_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)
    address_field.delete(0, END)

# function get data from GUI and write to Excel Spreadsheet
def insert():
    # if the entry field are left empty, then print 'empty input'
    # else, save the data entered by the user to the spreadsheet
    if (name_field.get() == '' and course_field.get()==''
    and sem_field.get() == '' and form_no_field.get()==''
    and contact_no_field.get() == '' and email_id_field.get()==''
    and address_field.get() == ''):
        print('Empty input')

    else:
        #assigning max_row and max_column values to variables
        current_row = sheet.max_row
        current_column = sheet.max_column

        # get() method returns data entered as string to
        # write at the right location in the spreadsheet
        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column= 2). value = course_field.get()
        sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
        sheet.cell(row=current_row + 1, column=4).value = form_no_field.get()
        sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
        sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
        sheet.cell(row=current_row + 1, column=7).value = address_field.get()

        # Save file with new content
        wb.save(r'/home/mr_unicorn/Documents/School/registration_Data.xlsx')

        # Return focus on the name_field
        name_field.focus_set()

        # Clear the form
        clear()

def customize():
    '''
    Implements the switching functionality between Dark and Light theme    
    ''''
    global theme_mode
    global dark
    global white
    global textFg

    if theme_mode:
        theme.configure(text= 'Dark', fg= dark, bg= white, activebackground= 'dark grey')
        register.configure(fg= dark, bg= white)
        root.configure(background= white)
        heading.configure(bg = white, fg= dark)
        name.configure(bg = white, fg = dark)
        course.configure(bg = white, fg= dark)
        sem.configure(bg = white, fg= dark)
        form_no.configure(bg = white, fg= dark)
        contact_no.configure(bg = white, fg= dark)
        email_id.configure(bg = white, fg= dark)
        address.configure(bg = white, fg= dark)
        theme_mode = False

    else:
        theme.configure(text= 'Light', fg= white, bg= dark)
        register.configure(fg= textFg, bg= dark)
        root.configure(background= dark)
        heading.configure(bg = dark, fg= textFg)
        name.configure(bg = dark, fg= textFg)
        course.configure(bg = dark, fg= textFg)
        sem.configure(bg = dark, fg= textFg)
        form_no.configure(bg = dark, fg= textFg)
        contact_no.configure(bg = dark, fg= textFg)
        email_id.configure(bg = dark, fg= textFg)
        address.configure(bg = dark, fg= textFg)
        theme_mode = True

# Driver Code containing the mainloop for tkinter window
if __name__== '__main__':

    # GUI main window
    root = Tk(className='Registration Form')
    root.configure(background= dark)
    root.geometry('550x350')

    excel()

    # Labels for different entry fields
    heading = Label(root, text = 'Form', bg = dark, fg=textFg)
    name = Label(root, text = 'Name', bg = dark, fg=textFg)
    course = Label(root, text = 'Course', bg = dark, fg= textFg)
    sem = Label(root, text = 'Semester', bg = dark, fg=textFg)
    form_no = Label(root, text = 'Form No.', bg = dark, fg=textFg)
    contact_no = Label(root, text ='Contact No.', bg = dark, fg=textFg)
    email_id = Label(root, text= 'Email id', bg = dark, fg=textFg)
    address = Label(root, text ='Address', bg = dark, fg=textFg)


    # placing widgets using grid()
    heading.grid(row=0, column=1)
    name.grid(row=1, column=0)
    course.grid(row=2, column=0)
    sem.grid(row=3, column=0)
    form_no.grid(row=4, column=0)
    contact_no.grid(row=5, column=0)
    email_id.grid(row=6, column=0)
    address.grid(row=7, column=0)

    # Creating entry fields for data input
    name_field = Entry(root)
    course_field = Entry(root)
    sem_field = Entry(root)
    form_no_field = Entry(root)
    contact_no_field = Entry(root)
    email_id_field = Entry(root)
    address_field = Entry(root)

    # Binding focus events to the Return key using the
    # bind() method
    name_field.bind('<Return>', focus1)
    course_field.bind('<Return>', focus2)
    sem_field.bind('<Return>', focus3)
    form_no_field.bind('<Return>', focus4)
    contact_no_field.bind('<Return>', focus5)
    email_id_field.bind('<Return>', focus6)


    # Placing the Entry widgets using grid() method
    name_field.grid(row=1, column=1, ipadx='100')
    course_field.grid(row=2, column=1, ipadx='100')
    sem_field.grid(row=3, column=1, ipadx="100")
    form_no_field.grid(row=4, column=1, ipadx="100")
    contact_no_field.grid(row=5, column=1, ipadx="100")
    email_id_field.grid(row=6, column=1, ipadx="100")
    address_field.grid(row=7, column=1, ipadx="100")

    # Calling excel function
    excel()

    # Register button for the Form
    register = Button(root, text= 'Register', fg= textFg,
    bg= dark, command=insert)
    register.grid(row=8, column=1)
    register.grid_configure(pady= 8)

    # Changing form theme
    theme = Button(root, text= 'Light', fg = white, bg= dark,
    activebackground = white ,command=customize)
    theme.config(anchor='sw')
    theme.place(x = 460, y= 270)


    # running the mainloop
    root.mainloop()





